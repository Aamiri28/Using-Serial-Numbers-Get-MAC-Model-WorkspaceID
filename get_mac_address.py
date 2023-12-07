from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoAlertPresentException
from selenium.webdriver.chrome.options import Options
import openpyxl
import time


web = webdriver.Chrome()
web.get('https://common.cloud.hpe.com/')

# Use WebDriverWait to wait for the element to be present
wait = WebDriverWait(web, 10)
user = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/main/div[2]/div/div/form/div[1]/div[2]/div[1]/div[2]/span/input")))

time.sleep(3)
UserName = "mohammed.zeeshan-a@hpe.com"
user.send_keys(UserName)

login = wait.until(EC.element_to_be_clickable((By.XPATH, "//html/body/div[5]/main/div[2]/div/div/form/div[2]/input")))
login.click()

time.sleep(20)

web.maximize_window()

login_account = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div/div[4]/div/div/div[3]/div/div[3]/button")))
web.execute_script("arguments[0].scrollIntoView();", login_account)
time.sleep(3)
login_account.click()

time.sleep(10)


manageccs = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[5]/div[3]/div[2]/div")))
web.execute_script("arguments[0].scrollIntoView();", manageccs)
time.sleep(2)
manageccs.click()

time.sleep(5)

devices = wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/nav/button[3]/div/div/span")))
web.execute_script("arguments[0].scrollIntoView();", devices)
time.sleep(2)
devices.click()

time.sleep(5)

"""
wait = WebDriverWait(web, 10)
SerialNumber = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div[2]/div[2]/div[3]/div/div[1]/div/div[1]/div/input")))
device = "CZJ32911PM"
SerialNumber.send_keys(device)

# Find the element containing the text you want to copy
time.sleep(2)

Serial = web.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div/table/tbody/tr/th/div/div/span")
Model = web.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div/table/tbody/tr/td[1]/div/span")
MAC = web.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div/table/tbody/tr/td[3]/div/span")
Folder = web.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div/table/tbody/tr/td[4]/div")
CustomerID = web.find_element(By.XPATH, "/html/body/div/div/div/div/div[3]/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div/table/tbody/tr/td[5]/div/span")


# Extract the text from the element

copied_serial = Serial.text
copied_model = Model.text
copied_mac = MAC.text
copied_folder = Folder.text
copied_customerid = CustomerID.text

# Create or open an Excel file using openpyxl
workbook = openpyxl.Workbook()
sheet = workbook.active

# Paste the text into a cell in Excel
sheet['A1'] = copied_serial
sheet['B1'] = copied_model 
sheet['C1'] = copied_mac
sheet['D1'] = copied_folder
sheet['E1'] = copied_customerid

# Save the Excel file
workbook.save('output.xlsx')

time.sleep(20)

"""
# Load the Excel file containing the serial numbers
workbook = openpyxl.load_workbook('serials.xlsx')
sheet = workbook.active

# Create or open the output Excel file
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet['A1'] = 'Serial Number'
output_sheet['B1'] = 'Model'
output_sheet['C1'] = 'MAC'
output_sheet['D1'] = 'Folder'
output_sheet['E1'] = 'Customer ID'

# Initialize variables outside the loop
copied_serial = ""
copied_model = ""
copied_mac = ""
copied_folder = ""
copied_customerid = ""

# Loop through each row in the Excel file (skipping the header row)
for row_num in range(2, sheet.max_row + 1):
    # Extract the serial number from the Excel file
    serial_number = sheet.cell(row=row_num, column=1).value

    # Enter the serial number in the web page
    
    SerialNumber = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/div[2]/div[3]/div/div[1]/div/div[1]/div/input")))
    SerialNumber.send_keys(Keys.CONTROL + "a")  # Select all the text in the input field
    SerialNumber.send_keys(Keys.BACKSPACE)  # Delete the selected text
    SerialNumber.send_keys(serial_number)
    time.sleep(3)

    try:
        # Extract the necessary information from the web page
        Serial = web.find_element(By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div[1]/table/tbody/tr/th/div/div/span")
        Model = web.find_element(By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div[1]/table/tbody/tr/td[1]/div/span")
        MAC = web.find_element(By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div[1]/table/tbody/tr/td[3]/div/span")
        Folder = web.find_element(By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div[1]/table/tbody/tr/td[4]/div/span")
        CustomerID = web.find_element(By.XPATH, "/html/body/div/div[1]/div/div/div/div/div/div/div[2]/div[2]/div[2]/div[3]/div/div[3]/div/div[1]/table/tbody/tr/td[5]/div/span")

        # Extract the text from the element

        copied_serial = Serial.text
        copied_model = Model.text
        copied_mac = MAC.text
        copied_folder = Folder.text
        copied_customerid = CustomerID.text

    except:
        # Handle the scenario when some details are missing
        Serial = ""
        copied_model = ""
        copied_mac = ""
        copied_folder = ""
        copied_customerid = ""

    # Append the details to the output Excel file
    output_row_num = output_sheet.max_row + 1
    output_sheet.cell(row=output_row_num, column=1, value=serial_number)
    output_sheet.cell(row=output_row_num, column=2, value=copied_model)
    output_sheet.cell(row=output_row_num, column=3, value=copied_mac)
    output_sheet.cell(row=output_row_num, column=4, value=copied_folder)
    output_sheet.cell(row=output_row_num, column=5, value=copied_customerid)

    # Save the output Excel file after processing each serial number
    output_workbook.save('output.xlsx')



